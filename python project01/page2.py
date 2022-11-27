from tkinter import *
from tkinter import messagebox
import ast
import pandas as pd
from openpyxl import load_workbook
import pandas as pd
import openpyxl
import os   
from datetime import timedelta, date
from openpyxl import load_workbook
from datetime import datetime



 




window=Tk()
window.title("SignUp")
window.geometry("1920x1080+300+200")
window.configure(bg="black")
window.resizable(True,True)

img=PhotoImage(file="login(00).png")
Label(window,image=img,bg="white").place(x=0,y=0)

frame=Frame(window,width=350,height=400,bg="white")

frame.place(x=600,y=260)

heading=Label(frame,text="SignUp" ,fg="#57a1f8",bg="white",font=('Montserrat',23,"bold"))

heading.place(x=100,y=5)

###############################

def signup():
    username=user.get()
    password=code.get()
    comform_password=conform_code.get()
    # wb=load_workbook("user_data.xlsx")
    
    
    if(password==conform_code):
        try:
            wb=load_workbook("user_data.xlsx")
            lb=openpyxl.Workbook()
            sheet1=lb.active
            
            
            
            lb.save("user_data.xlsx")

        except:
            lb=openpyxl.Workbook()
            sheet2=lb.active
            
            
            
            
            lb.save("user_data.xlsx")
            
            
            print("file creation done")  
        
        
        
def on_enter(e):
    user.delete(0,"end")
    
def on_leave(e):
    name=user.get()
    if(name==""):
        user.insert(0,"username")
        
    
    
    
user=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Montserrat",11,"bold"))
user.place(x=30,y=80)
user.insert(0,"username")

user.bind("<FocusIn>",on_enter)
user.bind("<FocusOut>",on_leave)

Frame(frame,width=295,height=2,bg="black").place(x=25,y=107)

########################

def on_enter(e):
    code.delete(0,"end")
    
def on_leave(e):
    name=code.get()
    if(name==""):
        code.insert(0,"password")

code=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Montserrat",11,"bold"))
code.place(x=30,y=150)
code.insert(0,"password")

code.bind("<FocusIn>",on_enter)
code.bind("<FocusOut>",on_leave)



Frame(frame,width=295,height=2,bg="black").place(x=25,y=177)



########################

def on_enter(e):
    conform_code.delete(0,"end")
    
def on_leave(e):
    name=conform_code.get()
    if(name==""):
        conform_code.insert(0,"confirm password")

conform_code=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Montserrat",11,"bold"))
conform_code.place(x=30,y=220)
conform_code.insert(0,"confirm password")

conform_code.bind("<FocusIn>",on_enter)
conform_code.bind("<FocusOut>",on_leave)



Frame(frame,width=295,height=2,bg="black").place(x=25,y=247)


#######################



Button(frame,width=39,pady=7,text="SignUp" ,font=("Montserrat",8,"bold"),fg="white",border=0,bg="#57a1f8",command=signup).place(x=25,y=288)


label=Label(frame,text="I have a account?" ,fg="black",bg="white",font=("Montserrat",9,"bold"))
label.place(x=75,y=354)

sign_up=Button(frame,width=6,text="Sign in" ,fg="#3498db",border=0,bg="white",cursor="hand2" ,font=("Montserrat",9,"bold"))
sign_up.place(x=215,y=354)
# 3498db
##############################
window.mainloop()